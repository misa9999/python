import openpyxl
from random import uniform

my_spreadsheet = openpyxl.Workbook()
my_spreadsheet.create_sheet("Spreadsheet1", 0)
my_spreadsheet.create_sheet("Spreadsheet2", 1)

spreadsheet1 = my_spreadsheet["Spreadsheet1"]
spreadsheet2 = my_spreadsheet["Spreadsheet2"]

for row in range(1, 11):
    request_num = row - 1
    price = round(uniform(10, 100), 2)
    spreadsheet1.cell(row, 1).value = request_num
    spreadsheet1.cell(row, 2).value = 1200 + row
    spreadsheet1.cell(row, 3).value = price

for row in range(1, 11):
    request_num = row - 1
    price = round(uniform(10, 100), 2)
    spreadsheet2.cell(row, 1).value = request_num
    spreadsheet2.cell(row, 2).value = 1200 + row
    spreadsheet2.cell(row, 3).value = price

my_spreadsheet.save("my_new_spreadsheet.xlsx")

# requests = openpyxl.load_workbook("excel/pedidos.xlsx")
# sheet_names = requests.sheetnames
# spreadsheet1 = requests["Página1"]

# for row in spreadsheet1["a1:c2"]:
#     for column in row:
#         print(column.value)

# for row in spreadsheet1:
#     if row[0].value is not None:
#         print(row[0].value, end=" ")
#     if row[1].value is not None:
#         print(row[1].value, end=" ")
#     if row[2].value is not None:
#         print(row[2].value)


# for row in range(5, 16):
#     request_num = row - 1
#     price = round(uniform(10, 100), 2)
#     spreadsheet1.cell(row, 1).value = request_num
#     spreadsheet1.cell(row, 2).value = 1200 + row
#     spreadsheet1.cell(row, 3).value = price


# requests.save("new_spreadsheet.xlsx")
